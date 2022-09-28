import os
import re
import xlrd

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter

wb = Workbook()
# grab the active worksheet
ws = wb.active
# Data can be assigned directly to cells
# ws['A1'] = 102.3
# Rows can also be appended
# ws.append([1, 2, 3])
# Save the file
# wb.save("sample.xlsx")

sheetNames = ['1 Idle', '2 CCC', '3 CVC', '4 Idle', '5 CCD', '6 Idle', '7 CCC', '8 Discharge']
for i in range(8):
    wb.create_sheet(sheetNames[i])
wb.remove(ws)  # remove the default work sheet
wb['8 Discharge']['B1'] = 'Discharge Capacity'

dir_list = os.listdir("Files")  # open folder containing the file

startCol = 1  # starting column
startRow = 2
for x in range(len(dir_list)):
    # print(x)
    file = open("Files/" + dir_list[x], encoding="ISO-8859-1")  # open file
    # file = open("Files/" + "_220817_box1_group5_001_7_2.TXT", encoding="ISO-8859-1")
    # if file.name.startswith("Files/_220817_box1_group5_001_7_2"):
        # print(00)
    if file.name.endswith(".TXT"):
        lines = file.readlines()  # load each line into array
        j = -1
        numbers = []  # 7 sub array for different tests, each subarray contain min, v, mA, mAh
        for i in range(len(lines)):  # every time the line contains "mJ" start a new block
            if "mJ" in lines[i]:
                j += 1
                numbers.append([])
            elif j != -1:
                filtered = list(
                    map(float,
                        re.findall("[0-9.]+(?![0-9]\))", lines[i])[1:]))  # extract the numbers in the line using regex
                if len(filtered) != 0:
                    numbers[j].append(filtered)

        for k in range(len(numbers)):
            if len(numbers) == 6:
                wb.active = wb[sheetNames[k + 1]]  # switch excel work book
            else:
                wb.active = wb[sheetNames[k]]  # switch excel work book
            ws = wb.active
            ws[get_column_letter(startCol) + '1'] = os.path.basename(file.name).split('.')[0]
            ws[get_column_letter(startCol)+'2'] = 'min'
            ws[get_column_letter(startCol+1) + '2'] = 'V'
            ws[get_column_letter(startCol+2) + '2'] = 'mA'
            ws[get_column_letter(startCol+3) + '2'] = 'mAh'
            for i in range(len(numbers[k])):
                for j in range(len(numbers[k][i])):
                    colLet = get_column_letter(startCol + j)
                    pos = colLet + str(i + 3)  # get the position in excel sheet
                    ws[pos] = numbers[k][i][j]

        if len(numbers)<7:
            print(file.name + "     # of data: " + str(len(numbers)))

        ws1 = xlrd.open_workbook("Files/" + dir_list[x].split(".")[0] + ".xls")
        wb['8 Discharge']['A'+str(startRow)] = dir_list[x].split(".")[0]
        wb['8 Discharge']['B'+str(startRow)] = ws1.sheet_by_index(0).cell_value(rowx=1, colx=5)

        startCol += 5
        startRow += 1

wb.save("sample.xlsx")

# for i in range(len(lines)):
